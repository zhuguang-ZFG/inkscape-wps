"""xlsx 导入/导出：列宽/行高 ↔ cell_w_mm/cell_h_mm 近似映射回归测试。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from inkscape_wps.core.office_export import export_xlsx
from inkscape_wps.core.office_import import import_xlsx_to_table_blob

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]


def _excel_col_width_to_mm(width: float) -> float:
    # 像素≈ (宽度*7)+5；像素(mm) = mm*96/25.4
    px = float(width) * 7.0 + 5.0
    return px * 25.4 / 96.0


def _excel_row_height_points_to_mm(height_points: float) -> float:
    return float(height_points) * 25.4 / 72.0


def _mm_to_excel_col_width(mm: float) -> float:
    px = float(mm) * 96.0 / 25.4
    return max(0.0, (px - 5.0) / 7.0)


def _mm_to_excel_row_height_points(mm: float) -> float:
    return float(mm) * 72.0 / 25.4


@unittest.skipIf(openpyxl is None, "openpyxl 未安装，跳过 xlsx 维度测试")
class TestOfficeXlsxDims(unittest.TestCase):
    def test_import_reads_column_row_dims(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions["A"].width = 10.0
        ws.row_dimensions[1].height = 20.0
        ws.cell(row=1, column=1).value = "X"

        with tempfile.TemporaryDirectory(prefix="inkscape-wps-xlsx-dims-") as td:
            p = Path(td) / "dim.xlsx"
            wb.save(str(p))
            blob = import_xlsx_to_table_blob(p, max_rows=1, max_cols=1)

        self.assertAlmostEqual(blob["cell_w_mm"], _excel_col_width_to_mm(10.0), delta=0.8)
        self.assertAlmostEqual(blob["cell_h_mm"], _excel_row_height_points_to_mm(20.0), delta=0.8)

    def test_export_sets_column_row_dims(self) -> None:
        table_blob = {
            "cell_w_mm": 28.0,
            "cell_h_mm": 12.0,
            "rows": 1,
            "cols": 1,
            "cells": [[{"text": "X", "html": None}]],
        }

        with tempfile.TemporaryDirectory(prefix="inkscape-wps-xlsx-dims-") as td:
            p = Path(td) / "out.xlsx"
            export_xlsx(p, table_blob=table_blob, prefer_soffice=False)

            wb = openpyxl.load_workbook(str(p), read_only=False, data_only=True)
            ws = wb.active
            col_w = ws.column_dimensions["A"].width
            row_h = ws.row_dimensions[1].height

        self.assertIsNotNone(col_w)
        self.assertIsNotNone(row_h)
        self.assertAlmostEqual(float(col_w), _mm_to_excel_col_width(28.0), delta=0.5)
        self.assertAlmostEqual(float(row_h), _mm_to_excel_row_height_points(12.0), delta=0.5)


if __name__ == "__main__":
    unittest.main()

